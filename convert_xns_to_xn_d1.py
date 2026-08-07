"""Convert the `xns` sheet of the consolidated workbook into a file that matches
the schema of ``data/xn_d1.csv`` (tab-separated, 11 columns).

Source : excel_format/bank_statement_analysis.xlsx  ->  sheet 'xns'
Output : excel_format/xns_xn_d1.csv

xn_d1 columns:
  cust_id, dr_cr_indctor, tran_date, prty_name, tran_amt_in_ac, tran_partclr,
  sal_flag, self_transfer, tran_type, category_of_txn, category_of_txn_l2

Mapping from xns (Sl.No, Date, Cheque No, Description, Amount, Category, Balance):
  dr_cr_indctor   <- sign of Amount ( < 0 -> 'D', else 'C' )
  tran_amt_in_ac  <- abs(Amount)
  tran_date       <- '01-Feb-26' -> '2026-02-01'
  prty_name       <- account holder Name (from 'Analysis' tab; --party override)
  cust_id         <- Account No. (from 'Analysis' tab; --cust-id override)
  tran_partclr    <- Description
  tran_type       <- inferred from narration prefix
  sal_flag        <- 1 if Category mentions 'salary'
  self_transfer   <- 1 if narration mentions OWN/SELF account
  category_of_txn / category_of_txn_l2 <- xns Category, verbatim in both

Run:  python convert_xns_to_xn_d1.py [--cust-id X] [--party "Name"]
"""
from __future__ import annotations

import argparse
import csv
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent
XLSX = BASE / "excel_format" / "bank_statement_analysis.xlsx"
OUT = BASE / "excel_format" / "xns_xn_d1.csv"

COLUMNS = [
    "cust_id", "dr_cr_indctor", "tran_date", "prty_name", "tran_amt_in_ac",
    "tran_partclr", "sal_flag", "self_transfer", "tran_type",
    "category_of_txn", "category_of_txn_l2",
]

SELF_KEYWORDS = ("OWN A/C", "OWN ACCOUNT", " OWN ", "/OWN ", "SELF")


def infer_tran_type(desc: str) -> str:
    d = (desc or "").upper()
    if d.startswith("UPI") or "UPI/" in d:
        return "UPI"
    if d.startswith("MB:"):
        return "MB"
    if "IMPS" in d:
        return "IMPS"
    if "NACH" in d:
        return "NACH"
    if d.startswith("NEFT") or " NEFT " in d:
        return "NEFT"
    if d.startswith(("ATL/", "ATI/", "ATW")) or "ATM" in d:
        return "ATM"
    return ""


def to_iso_date(raw: str) -> str:
    """'01-Feb-26' -> '2026-02-01'. Returns '' if unparseable."""
    raw = (raw or "").strip()
    for fmt in ("%d-%b-%y", "%d-%b-%Y"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return raw


def fmt_amount(amt) -> str:
    a = abs(amt)
    return str(int(a)) if float(a).is_integer() else str(a)


def read_account_info(wb):
    """Pull Name and Account No. from the 'Analysis' tab's Account Info block."""
    name = acct = ""
    if "Analysis" in wb.sheetnames:
        for row in wb["Analysis"].iter_rows(values_only=True):
            key = str(row[0]).strip() if row and row[0] else ""
            val = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if key == "Name":
                name = val
            elif key == "Account No.":
                acct = val
    return name, acct


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--cust-id", default=None, help="override cust_id (default: Account No. from Analysis)")
    ap.add_argument("--party", default=None, help="override prty_name (default: Name from Analysis)")
    args = ap.parse_args()

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    if "xns" not in wb.sheetnames:
        raise SystemExit(f"'xns' sheet not found in {XLSX}")

    name, acct = read_account_info(wb)
    cust_id = args.cust_id or acct or ""
    party = args.party or name or ""

    rows = list(wb["xns"].iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(header)}

    out_rows, skipped = [], 0
    for r in rows[1:]:
        amount = r[idx["Amount"]]
        date = r[idx["Date"]]
        if amount is None or not date:
            skipped += 1
            continue
        desc = str(r[idx["Description"]] or "")
        category = str(r[idx["Category"]] or "")
        out_rows.append([
            cust_id,
            "D" if float(amount) < 0 else "C",
            to_iso_date(str(date)),
            party,
            fmt_amount(amount),
            desc,
            "1" if "salary" in category.lower() else "",
            "1" if any(k in desc.upper() for k in SELF_KEYWORDS) else "0",
            infer_tran_type(desc),
            category,
            category,
        ])

    with open(OUT, "w", newline="") as fh:
        w = csv.writer(fh, delimiter="\t")
        w.writerow(COLUMNS)
        w.writerows(out_rows)

    print(f"cust_id={cust_id!r}  prty_name={party!r}")
    print(f"Wrote {len(out_rows)} rows (skipped {skipped}) -> {OUT}")


if __name__ == "__main__":
    main()
