"""Ingest an *xns* statement file into an ``xn_d1.csv``-compatible per-customer
dataset (tab-separated), carrying the running Balance across as ``eod_balance``.

The canonical ``data/xn_d1.csv`` is never touched. Output goes to
``data/xns_generated/<cust>_xn_d1.csv`` so the pipeline can be pointed at it
(via ``data.loader.load_transactions(path=...)``) without any downstream change.

Accepted input:
  * an .xlsx workbook containing an ``xns`` sheet (else the first sheet), or
  * a .csv / .tsv with the xns headers
    (Sl.No, Date, Cheque No, Description, Amount, Category, Balance).

Output columns (xn_d1 schema + balance):
  cust_id, dr_cr_indctor, tran_date, prty_name, tran_amt_in_ac, tran_partclr,
  sal_flag, self_transfer, tran_type, category_of_txn, category_of_txn_l2,
  eod_balance

Run:  python -m tools.xns_ingest <xns_file> --cust-id <id> [--party "Name"]
"""
from __future__ import annotations

import argparse
import csv
import re
from datetime import date, datetime
from pathlib import Path

from tools.excel_utils import clean_number

BASE = Path(__file__).resolve().parent.parent
OUT_DIR = BASE / "data" / "xns_generated"

OUT_COLUMNS = [
    "cust_id", "dr_cr_indctor", "tran_date", "prty_name", "tran_amt_in_ac",
    "tran_partclr", "sal_flag", "self_transfer", "tran_type",
    "category_of_txn", "category_of_txn_l2", "eod_balance",
]

SELF_KEYWORDS = ("OWN A/C", "OWN ACCOUNT", " OWN ", "/OWN ", "SELF")

# canonical field -> substrings that identify its column header (case-insensitive)
_FIELD_MATCHERS = [
    ("desc", ("description", "narration", "particular", "remarks")),
    ("amount", ("amount", "amt")),
    ("balance", ("balance", "closing bal")),
    ("date", ("txn date", "transaction date", "value date", "tran date", "date")),
    ("category", ("category",)),
    ("cheque", ("cheque", "chq")),
    ("acct_no", ("account no", "account number", "a/c no", "acct no")),
]


def _match_field(label: str):
    """Map a header cell to a canonical field via substring match, or None."""
    l = str(label or "").strip().lower()
    if not l:
        return None
    for key, needles in _FIELD_MATCHERS:
        if any(n in l for n in needles):
            return key
    return None


def infer_tran_type(desc: str) -> str:
    d = (desc or "").upper()
    if d.startswith("UPI") or "UPI/" in d:
        return "UPI"
    if d.startswith("MB:"):
        return "MB"
    if "IMPS" in d:
        return "IMPS"
    if "NACH" in d:
        return "NACH"
    if d.startswith("NEFT") or " NEFT " in d:
        return "NEFT"
    if d.startswith(("ATL/", "ATI/", "ATW")) or "ATM" in d:
        return "ATM"
    return ""


def to_iso_date(raw) -> str:
    """Normalise a date to 'YYYY-MM-DD'.

    Handles text ('01-Feb-26', '01/02/2026'), real datetime/date cells, and
    datetime strings carrying a time part ('2026-02-01 00:00:00'). Indian
    statements are day-first. Returns the input unchanged if unparseable.
    """
    if raw is None:
        return ""
    if isinstance(raw, (datetime, date)):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    if not s:
        return ""
    # drop a trailing time component ('2026-02-01 00:00:00' / '2026-02-01T..')
    s = re.split(r"[ T]", s, 1)[0] if re.match(r"^\d{4}-\d{2}-\d{2}[ T]", s) else s
    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # last-resort: pandas' flexible parser (day-first for Indian statements)
    try:
        import pandas as pd
        ts = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.notna(ts):
            return ts.strftime("%Y-%m-%d")
    except Exception:
        pass
    return s


def fmt_amount(amt) -> str:
    a = abs(amt)
    return str(int(a)) if float(a).is_integer() else str(a)


def _fmt_cell(c):
    if c is None:
        return ""
    if isinstance(c, (datetime, date)):     # real Excel date cell
        return c.strftime("%Y-%m-%d")
    return str(c).strip()


def _locate_columns(rows):
    """Find the header row and map canonical field -> column index.

    A header row is one that has a Description, an Amount, and a Date column
    (matched by substring, so extra columns and header variants are tolerated).
    """
    for i, r in enumerate(rows):
        idx = {}
        for j, cellv in enumerate(r):
            key = _match_field(cellv)
            if key and key not in idx:
                idx[key] = j
        if "desc" in idx and "amount" in idx and "date" in idx:
            return i, idx
    raise ValueError(
        "Could not locate an xns header row (need Description + Amount + Date). "
        "Check the file has a header row with those columns."
    )


def _read_table(xns_path: Path):
    """Return (rows, account_name, account_no) from an xlsx or csv xns file.

    Delimiter is auto-detected for text files by trying tab/comma/semicolon/pipe
    and keeping the one that yields a locatable header (robust to commas inside
    values, e.g. 'ICICI Bank, India' or '(11,16,370.00)'). For workbooks, every
    sheet is scanned for the ledger.
    """
    import io
    name = acct = ""

    if xns_path.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(xns_path, read_only=True, data_only=True)
        if "Analysis" in wb.sheetnames:
            for r in wb["Analysis"].iter_rows(values_only=True):
                k = str(r[0]).strip() if r and r[0] else ""
                v = str(r[1]).strip() if len(r) > 1 and r[1] else ""
                if k == "Name":
                    name = v
                elif k == "Account No.":
                    acct = v
        order = (["xns"] if "xns" in wb.sheetnames else []) + list(wb.sheetnames)
        fallback = None
        for sheet in order:
            rows = [[_fmt_cell(c) for c in r] for r in wb[sheet].iter_rows(values_only=True)]
            fallback = fallback or rows
            try:
                _locate_columns(rows)
                return rows, name, acct
            except ValueError:
                continue
        return fallback or [], name, acct

    # text file: try delimiters, keep the one that locates a header
    with open(xns_path, newline="", encoding="utf-8-sig", errors="replace") as fh:
        text = fh.read()
    best = None
    for delim in ("\t", ",", ";", "|"):
        rows = [[c.strip() for c in row] for row in csv.reader(io.StringIO(text), delimiter=delim)]
        try:
            _locate_columns(rows)
            return rows, name, acct
        except ValueError:
            if best is None or max((len(r) for r in rows), default=0) > max((len(r) for r in best), default=0):
                best = rows
    return best or [], name, acct


def ingest_xns(xns_path, cust_id=None, party=None) -> Path:
    """Convert an xns file to an xn_d1-compatible CSV. Returns the output path."""
    xns_path = Path(xns_path)
    rows, name, acct = _read_table(xns_path)
    hdr, idx = _locate_columns(rows)

    def get(r, key):
        j = idx.get(key)
        return r[j].strip() if j is not None and j < len(r) else ""

    # cust_id precedence: explicit arg > Analysis-tab account > Account No. column > filename
    acct_col = ""
    if "acct_no" in idx:
        for r in rows[hdr + 1:]:
            v = get(r, "acct_no")
            if v:
                acct_col = v
                break
    cid = str(cust_id or acct or acct_col or xns_path.stem)
    pty = party or name or ""

    out_rows = []
    for r in rows[hdr + 1:]:
        raw_amt = get(r, "amount")
        date = get(r, "date")
        amount = clean_number(raw_amt)
        if amount is None or not isinstance(amount, (int, float)) or not date:
            continue
        desc = get(r, "desc")
        category = get(r, "category")
        balance = clean_number(get(r, "balance"))
        out_rows.append([
            cid,
            "D" if float(amount) < 0 else "C",
            to_iso_date(date),
            pty,
            fmt_amount(amount),
            desc,
            "1" if "salary" in category.lower() else "",
            "1" if any(k in desc.upper() for k in SELF_KEYWORDS) else "0",
            infer_tran_type(desc),
            category,
            category,
            "" if balance is None else balance,
        ])

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / f"{cid}_xn_d1.csv"
    with open(out_path, "w", newline="") as fh:
        w = csv.writer(fh, delimiter="\t")
        w.writerow(OUT_COLUMNS)
        w.writerows(out_rows)
    return out_path


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xns_file", help="path to an xns .xlsx or .csv/.tsv")
    ap.add_argument("--cust-id", default=None)
    ap.add_argument("--party", default=None)
    args = ap.parse_args()

    out = ingest_xns(args.xns_file, cust_id=args.cust_id, party=args.party)
    with open(out) as fh:
        n = sum(1 for _ in fh) - 1
    print(f"Wrote {n} rows -> {out}")


if __name__ == "__main__":
    main()
