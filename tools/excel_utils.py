"""Shared helpers for building .xlsx workbooks from tabular data.

Used by both `build_excel_format.py` (sample CSV -> workbook) and
`tools/statement_workbook.py` (per-customer computed workbook).
"""
from __future__ import annotations

import csv
import re
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter

_PREFIX = re.compile(r"^sheet\d+(?:_\d+)?_")


def read_rows(path) -> list[list[str]]:
    """Read a tab-separated file into stripped rows (handles multi-line quoted cells)."""
    with open(path, newline="") as fh:
        return [[c.strip() for c in row] for row in csv.reader(fh, delimiter="\t")]


def cell(row: list[str], idx: int) -> str:
    return row[idx].strip() if idx < len(row) else ""


def clean_number(val):
    """'(1,786.00)' -> -1786.0 ; '3,00,830.86 ' -> 300830.86 ; '' -> None ; else original."""
    if val is None:
        return None
    s = str(val).strip()
    if s == "":
        return None
    neg = s.startswith("(") and s.endswith(")")
    body = s[1:-1] if neg else s
    body = body.replace(",", "").strip()
    if re.fullmatch(r"-?\d+(?:\.\d+)?", body):
        num = float(body)
        if neg:
            num = -num
        return int(num) if num.is_integer() else num
    return s


def tab_name(filename: str) -> str:
    """'sheet2_1_Loan_credit_Recieved.csv' -> 'Loan_credit_Recieved' (<=31 chars)."""
    name = _PREFIX.sub("", Path(filename).stem)
    return name[:31]


def autofit(ws, frames: list[pd.DataFrame]) -> None:
    widths: dict[int, int] = {}
    for df in frames:
        for j, col in enumerate(df.columns, start=1):
            longest = max([len(str(col))] + [len(str(v)) for v in df[col].tolist()], default=10)
            widths[j] = min(60, max(widths.get(j, 0), longest + 2))
    for j, w in widths.items():
        ws.column_dimensions[get_column_letter(j)].width = w


def write_sheet(writer, sheet: str, payload) -> None:
    """payload is a DataFrame, or a list of (title, DataFrame) stacked with a gap."""
    if isinstance(payload, pd.DataFrame):
        payload.to_excel(writer, sheet_name=sheet, index=False)
        autofit(writer.book[sheet], [payload])
        return
    row = 0
    frames = []
    for title, df in payload:
        ws = writer.book.create_sheet(sheet) if sheet not in writer.book.sheetnames else writer.book[sheet]
        ws.cell(row=row + 1, column=1, value=title)
        df.to_excel(writer, sheet_name=sheet, index=False, startrow=row + 1)
        frames.append(df)
        row += len(df) + 3
    autofit(writer.book[sheet], frames)
